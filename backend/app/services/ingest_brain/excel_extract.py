"""F09 EXTRACT — messy Excel → list[RawTable] (Phase 1).

Goes beyond ``ingest/excel_reader.py`` (which detects a single header row and
clean-reads one block per sheet) to handle the cases that silently corrupt data
today, per the F09 design §5:

  * merged cells            → forward-fill the top-left value into covered cells
  * N tables in one sheet   → region-detect by blank-row/blank-col runs → split
  * 2-row hierarchical head → flatten ``parent · child`` (parent span fwd-filled)
  * title/banner rows       → skipped via header scoring
  * shifted / stray rows    → flagged in notes, never silently merged

Pure openpyxl + stdlib. NO pandas dependency here (we read raw cell grids so we
control merge handling). NEVER raises — any failure returns [] and the caller
falls back to ``excel_reader.read_excel``.

Public:
    extract_tables(path, *, filename="") -> list[RawTable]
"""
from __future__ import annotations

import logging
import re
from typing import Any, List, Optional, Tuple

from app.services.ingest_brain.contract import RawTable

logger = logging.getLogger(__name__)

_NULL_TOKENS = {"", "na", "n/a", "null", "none", "-", "?", "nan"}
_MAX_SCAN_HEADER = 8          # how many top rows of a region to consider as header
_MIN_REGION_ROWS = 2          # header + ≥1 data row


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    s = str(v).strip().lower()
    return s in _NULL_TOKENS


def _is_label(v: Any) -> bool:
    """Non-blank, non-numeric-looking string → a header label candidate."""
    if _is_blank(v):
        return False
    s = str(v).strip()
    try:
        float(s.replace(",", ""))
        return False
    except (ValueError, TypeError):
        return True


def _col_letter(idx0: int) -> str:
    """0-based column index → Excel letter (0→A)."""
    s, n = "", idx0 + 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


# --- load + merged-cell forward fill ----------------------------------------
def _load_grid(path: str) -> List[Tuple[str, List[List[Any]]]]:
    """Return [(sheet_name, grid)] with merged cells forward-filled.

    grid[r][c] = value. Merged ranges have their top-left value copied into every
    covered cell so region/label detection sees real values, not nulls.
    """
    import openpyxl
    out: List[Tuple[str, List[List[Any]]]] = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            width = max((len(r) for r in rows), default=0)
            grid: List[List[Any]] = [list(r) + [None] * (width - len(r)) for r in rows]
            # forward-fill merged ranges (top-left value into the whole range)
            for mc in list(ws.merged_cells.ranges):
                r0, r1 = mc.min_row - 1, mc.max_row - 1
                c0, c1 = mc.min_col - 1, mc.max_col - 1
                if r0 < 0 or c0 < 0 or r0 >= len(grid):
                    continue
                val = grid[r0][c0] if c0 < len(grid[r0]) else None
                for r in range(r0, min(r1, len(grid) - 1) + 1):
                    for c in range(c0, min(c1, width - 1) + 1):
                        grid[r][c] = val
            out.append((ws.title, grid))
    finally:
        wb.close()
    return out


# --- region detection (N tables per sheet) ----------------------------------
def _row_blank(grid: List[List[Any]], r: int) -> bool:
    return all(_is_blank(v) for v in grid[r])


def _split_row_regions(grid: List[List[Any]]) -> List[Tuple[int, int]]:
    """Split a sheet grid into (start,end) row bands separated by blank-row runs."""
    bands: List[Tuple[int, int]] = []
    start: Optional[int] = None
    for r in range(len(grid)):
        if _row_blank(grid, r):
            if start is not None:
                bands.append((start, r - 1))
                start = None
        else:
            if start is None:
                start = r
    if start is not None:
        bands.append((start, len(grid) - 1))
    return bands


def _trim_cols(block: List[List[Any]]) -> List[List[Any]]:
    """Drop fully-blank leading/trailing columns of a block."""
    if not block:
        return block
    width = max(len(r) for r in block)
    block = [r + [None] * (width - len(r)) for r in block]
    keep = [c for c in range(width) if any(not _is_blank(r[c]) for r in block)]
    if not keep:
        return []
    return [[r[c] for c in keep] for r in block]


# --- header detection + hierarchical flatten --------------------------------
def _label_coverage(row: List[Any]) -> float:
    non_null = [v for v in row if not _is_blank(v)]
    if not non_null:
        return 0.0
    return sum(1 for v in non_null if _is_label(v)) / len(non_null)


def _has_adjacent_dups(row: List[Any]) -> bool:
    """A grouping/parent header row often has repeated adjacent values (the trace
    of a merged cell spanning several columns, e.g. 2024 | 2024 | 2025)."""
    for k in range(len(row) - 1):
        if not _is_blank(row[k]) and row[k] == row[k + 1]:
            return True
    return False


def _find_header_rows(block: List[List[Any]]) -> List[int]:
    """Return the header row indices within a block (1 or 2 rows).

    Picks the first row with ≥60% labels spanning ≥60% of columns (the *child*
    header). Then looks for a hierarchical PARENT:
      - upward: a spanning row just above that is a grouping row — adjacent
        duplicate values (merged-cell trace) OR low label-coverage but wide
        (e.g. years ``2024 | 2024 | 2025``) → ``[i-1, i]``.
      - downward: a 2nd label row below with data under it → ``[i, i+1]``.
    Upward parents (the merged-year/quarter case) take priority.
    """
    width = max((len(r) for r in block), default=0)
    min_cover = max(2, int(0.6 * width))
    n = len(block)
    for i in range(min(n, _MAX_SCAN_HEADER)):
        row = block[i]
        non_null = [v for v in row if not _is_blank(v)]
        if len(non_null) < min_cover:
            continue
        if _label_coverage(row) < 0.6:
            continue
        # upward parent (grouping row above the label row)
        if i - 1 >= 0:
            prev = block[i - 1]
            prev_nn = [v for v in prev if not _is_blank(v)]
            if len(prev_nn) >= min_cover and (_has_adjacent_dups(prev) or _label_coverage(prev) < 0.6):
                return [i - 1, i]
        # downward second header
        if i + 2 < n:
            nxt, after = block[i + 1], block[i + 2]
            if (_label_coverage(nxt) >= 0.6 and len([v for v in nxt if not _is_blank(v)]) >= min_cover
                    and _label_coverage(after) < 0.6):
                return [i, i + 1]
        return [i]
    return [0]


def _norm_name(v: Any) -> str:
    s = re.sub(r"\s+", " ", str(v).strip()) if not _is_blank(v) else ""
    return s


def _flatten_header(block: List[List[Any]], hdr_idx: List[int]) -> List[str]:
    """Build flat unique column names from 1 or 2 header rows.

    2-row: forward-fill the parent across its span, then ``parent · child``.
    """
    width = max(len(r) for r in block)
    if len(hdr_idx) == 1:
        top = block[hdr_idx[0]] + [None] * (width - len(block[hdr_idx[0]]))
        names = [_norm_name(v) or f"col_{_col_letter(c)}" for c, v in enumerate(top)]
    else:
        top = list(block[hdr_idx[0]]) + [None] * (width - len(block[hdr_idx[0]]))
        bot = list(block[hdr_idx[1]]) + [None] * (width - len(block[hdr_idx[1]]))
        # forward-fill parent across blanks (covers un-merged repeated parents)
        filled, last = [], ""
        for v in top:
            nv = _norm_name(v)
            if nv:
                last = nv
            filled.append(last)
        names = []
        for c in range(width):
            parent, child = filled[c], _norm_name(bot[c])
            if parent and child:
                names.append(f"{parent} · {child}")
            elif child:
                names.append(child)
            elif parent:
                names.append(parent)
            else:
                names.append(f"col_{_col_letter(c)}")
    # de-dup
    seen, out = {}, []
    for nm in names:
        if nm in seen:
            seen[nm] += 1
            out.append(f"{nm}_{seen[nm]}")
        else:
            seen[nm] = 0
            out.append(nm)
    return out


def _block_to_table(block: List[List[Any]], sheet: str, filename: str,
                    band: Tuple[int, int]) -> Optional[RawTable]:
    block = _trim_cols(block)
    if len(block) < _MIN_REGION_ROWS:
        return None
    hdr_idx = _find_header_rows(block)
    header = _flatten_header(block, hdr_idx)
    width = len(header)
    notes: List[str] = []
    data_start = hdr_idx[-1] + 1
    rows: List[List[Any]] = []
    dropped_blank = 0
    for r in range(data_start, len(block)):
        row = list(block[r]) + [None] * (width - len(block[r]))
        row = row[:width]
        if all(_is_blank(v) for v in row):
            dropped_blank += 1
            continue
        filled = sum(1 for v in row if not _is_blank(v))
        if filled < max(1, int(0.25 * width)):
            notes.append(f"row {r + band[0] + 1}: sparse ({filled}/{width} filled) — kept, check if a note row")
        rows.append(row)
    if not rows:
        return None
    if len(hdr_idx) == 2:
        notes.insert(0, "flattened a 2-row hierarchical header (parent · child)")
    if dropped_blank:
        notes.insert(0, f"dropped {dropped_blank} fully-blank row(s)")
    bbox = f"{_col_letter(0)}{band[0] + 1}:{_col_letter(width - 1)}{band[1] + 1}"
    name = re.sub(r"[^a-z0-9_]+", "_", sheet.strip().lower()).strip("_")[:48] or "table"
    return RawTable(name=name, header=header, rows=rows, source_file=filename,
                    sheet=sheet, region_bbox=bbox, notes=notes)


def extract_tables(path: str, *, filename: str = "") -> List[RawTable]:
    """Messy Excel → clean RawTables. NEVER raises (returns [] on failure)."""
    try:
        out: List[RawTable] = []
        merged_seen = False
        for sheet, grid in _load_grid(path):
            bands = _split_row_regions(grid)
            multi = len([b for b in bands if b[1] - b[0] + 1 >= _MIN_REGION_ROWS]) > 1
            for band in bands:
                block = [grid[r] for r in range(band[0], band[1] + 1)]
                tbl = _block_to_table(block, sheet, filename, band)
                if tbl is None:
                    continue
                if multi:
                    tbl.notes.insert(0, f"1 of multiple tables found in sheet '{sheet}'")
                out.append(tbl)
        return out
    except Exception:  # noqa: BLE001 — fail-soft, caller falls back
        logger.exception("ingest_brain.extract_tables failed for %s", path)
        return []
